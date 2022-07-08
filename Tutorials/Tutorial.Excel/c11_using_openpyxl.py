from openpyxl import load_workbook, Workbook

wb = load_workbook('test.xlsx')
ws = wb.active

print('\nUsing break')
for num, row in enumerate(ws.rows):
    if num >= 2:
        break
    print(row)

print('\nUsing slice')
for row in ws['A1:G2']:
    print(row)

print('\nCheck dimensions')
print(f'Dimensions: {ws.dimensions}')
print(f'Max Row: {ws.max_row}')
print(f'Max Column: {ws.max_column}')
