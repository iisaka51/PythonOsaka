import openpyxl
from pathlib import Path

class ExcelSpilter(object):
    def __init__(self, worksheet, min_row=1, max_row=100, column, delimitor):
        self.worksheet=worksheet
        self.min_row=min_row
        self.max_row=max_row
        self.column=column
        self.delimitor=delimitor
        self.workbook = None
        self.newbook = None
        self.ws = None
        self.data = list()

    def read(self, path):
        self.wb = self.load_workbook(path)
        self.s = self.load_workbook(self.worksheet)

    def list_sheet(self, path):
        self.read(path)
        for sheet in self.wb:
            print(sheet)


# ブック、シートを開く
wb = openpyxl.load_workbook("請求書_202106.xlsx")
ws = wb["Sheet1"]

# A3セルに宛名の書き込み
ws.cell(3, 1).value = "経済産業省"

# ファイル名を指定してブックを保存
wb.save("請求書_202106_signed.xlsx")
