from openpyxl import load_workbook, Workbook

wb = Workbook()
ws = wb.active

data = (10, 20, 30, 40)

ws.append(data)
ws['A2'] = '=SUM(A1:D1)'

wb.save('output_formula.xlsx')

wb2 = load_workbook('output_formula.xlsx')
ws2 = wb2.active

wb3 = load_workbook('output_formula.xlsx', data_only=True)
ws3 = wb3.active

# ws2['A2'].value
# ws3['A2'].value
