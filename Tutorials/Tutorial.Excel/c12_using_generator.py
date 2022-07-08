from openpyxl import load_workbook, Workbook

wb = load_workbook('test.xlsx')
ws = wb.active
ws2 = wb['Sheet1']

for row in ws.iter_rows(min_row=1, max_col=7, max_row=2):
    for cell in row:
        print(cell, cell.value)

assert ws == ws2

# wb.active
# ws['A1:G2']
