from openpyxl import load_workbook, Workbook
from dataclasses import dataclass
from datetime import datetime

@dataclass
class StockPrice:
    date: datetime
    hight: int
    low: int
    open: int
    close: int
    volume: int
    adjClose: int

    def value(self):
        return (self.date, self.hight, self.open, self.close,
                self.volume, self.adjClose )


wb = load_workbook('test.xlsx')
ws = wb.active

data = list()
columns_names = ws.iter_rows(min_row=1, max_row=1,
                             max_col=7, values_only=True).__next__()

values = ws.iter_rows(min_row=2, max_col=7, values_only=True)
for val in values:
    if val[6] < 100:
        data.append(StockPrice(*val))

# data[-1]
# data[-1].value()
# data[-1].adjClose
