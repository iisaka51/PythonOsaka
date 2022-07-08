from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
ws = wb.active

for values in ws.iter_rows(min_row=1, max_col=7, max_row=2,
                        values_only=True):
    print(values)
