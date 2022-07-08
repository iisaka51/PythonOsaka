from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference

wb = load_workbook('test.xlsx')
ws = wb.active

chart = LineChart()
chart.width = 20    # default is 15
chart.height = 15   # default is 7
chart.title = 'TSLA StockChart'
chart.x_axis.title = 'Date'
chart.y_axis.title = 'Adj Close Price'
chart.legend.position = 'b'

data = Reference(worksheet=ws,
                min_row=1, max_row=ws.max_row,
                min_col=ws.max_column, max_col=ws.max_column)
category = Reference(worksheet=ws,
                min_row=2, max_row=ws.max_row,
                min_col=1, max_col=1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(category)

ws.add_chart(chart, 'I2')
wb.save('TSLA.xlsx')
