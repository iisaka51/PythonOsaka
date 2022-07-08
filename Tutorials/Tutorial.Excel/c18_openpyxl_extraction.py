from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
ws = wb.active

data = list()
column_names = ws.iter_rows(min_row=1, max_row=1,
                            max_col=7, values_only=True).__next__()
data.append(column_names)

values = ws.iter_rows(min_row=2, max_col=7, values_only=True)
for val in values:
    if val[6] < 100:
        data.append(val)

# len(data)
# data[0]
# data[-1]
