from openpyxl import load_workbook, Workbook
from dataclasses import dataclass
from datetime import datetime

wb = load_workbook('test.xlsx')
ws = wb.active

stock_data = list()
columns_names = ws.iter_rows(min_row=1, max_row=1,
                             max_col=7, values_only=True).__next__()

values = ws.iter_rows(min_row=2, max_col=7, values_only=True)
for val in values:
    if val[6] < 100:
        stock_data.append(val)

new_wb = Workbook()
new_ws = new_wb.active

for data in stock_data:
    new_ws.append(data)

new_wb.save('output2.xlsx')

wb2 = load_workbook('output2.xlsx')
ws2 = wb2.active

# ws2.dimensions
# ws2.max_row
# ws2.max_column
