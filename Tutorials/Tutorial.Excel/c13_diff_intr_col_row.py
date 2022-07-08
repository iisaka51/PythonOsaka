from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
ws = wb.active
ws.title = 'TSLA'

print('\nusing iter_rows()')
for row in ws.iter_rows(min_row=1, max_row=2,
                        min_col=1, max_col=7):
    print(row)

print('\nusing iter_cols()')
for column in ws.iter_cols(min_row=1, max_row=2,
                           min_col=1, max_col=7):
    print(column)
