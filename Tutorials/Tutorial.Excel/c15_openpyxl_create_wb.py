from openpyxl import Workbook

wb = Workbook()

names1 = wb.sheetnames

ws = wb.active
v1 = ws.title

ws = wb.create_sheet("TSLA")     # 最後に追加 (default)
v2 = ws.title

names2 = wb.sheetnames
ws = wb.create_sheet("APPL",0)   # 指定位置に挿入
names3 = wb.sheetnames
ws = wb.create_sheet("MSFT",-1)  # 末尾からの指定位置に挿入
names4 = wb.sheetnames
del wb['Sheet']
names5 = wb.sheetnames


# names1
# v1, v2
# names2
# names3
# names4
# names5
